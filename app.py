import streamlit as st
import gspread
import json
from datetime import datetime
import pytz  
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- MATIKAN WARNING & FITUR CHAINED ASSIGNMENT SECARA TOTAL ---
import warnings
warnings.filterwarnings('ignore', category=FutureWarning)
pd.options.mode.chained_assignment = None

# --- KONFIGURASI HALAMAN WEB ---
st.set_page_config(page_title="Office Attendance Cloud", page_icon="🏢", layout="centered")

# ==================== SISTEM KEAMANAN LOGIN ====================
def cek_login():
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False

    if not st.session_state["authenticated"]:
        st.subheader("🔒 Akses Terbatas - Absensi Kantor")
        password = st.text_input("Masukkan Password Akses:", type="password")
        if st.button("Masuk", type="primary", use_container_width=True):
            if password == "kantor123": 
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("❌ Password salah!")
        return False
    return True

if cek_login():
    # ==================== KONEKSI GOOGLE SHEETS VIA GSPREAD ====================
    def dapatkan_koneksi_spreadsheet():
        try:
            kredensial_mentah = st.secrets["connections"]["gsheets"]["service_account"]
            spreadsheet_url = st.secrets["connections"]["gsheets"]["spreadsheet"]
            
            if isinstance(kredensial_mentah, str):
                kredensial_mentah = kredensial_mentah.replace('\\n', '\n')
                kredensial_json = json.loads(kredensial_mentah, strict=False)
            else:
                kredensial_json = dict(kredensial_mentah)
            
            gc = gspread.service_account_from_dict(kredensial_json)
            sh = gc.open_by_url(spreadsheet_url)
            return sh
        except Exception as e:
            st.error(f"❌ Masalah Koneksi Google Sheets: {e}")
            st.stop()

    # ==================== FUNGSI OTOMATISASI VISUAL SHEETS ====================
    def perbarui_desain_visual_sheet(ws, jumlah_kolom=8):
        try:
            ws.columns_auto_resize(1, jumlah_kolom)
            total_baris = len(ws.get_all_values())
            if total_baris == 0:
                return

            requests = []
            
            # Format Header (Baris 1) - Biru Eksekutif
            requests.append({
                "repeatCell": {
                    "range": {"sheetId": ws.id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": jumlah_kolom},
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": {"red": 0.1, "green": 0.4, "blue": 0.8},
                            "textFormat": {"foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}, "bold": True},
                            "horizontalAlignment": "CENTER"
                        }
                    },
                    "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)"
                }
            })
            
            # Format Efek Zebra
            for i in range(1, total_baris):
                warna_latar = {"red": 0.92, "green": 0.96, "blue": 1.0} if i % 2 == 1 else {"red": 1.0, "green": 1.0, "blue": 1.0}
                requests.append({
                    "repeatCell": {
                        "range": {"sheetId": ws.id, "startRowIndex": i, "endRowIndex": i + 1, "startColumnIndex": 0, "endColumnIndex": jumlah_kolom},
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": warna_latar,
                                "horizontalAlignment": "CENTER"
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment)"
                    }
                })
            
            ws.spreadsheet.batch_update({"requests": requests})
        except Exception:
            pass

    # --- REUSABLE EXCEL EXPORT FUNCTION ---
    def buat_excel_bytes(df, sheet_name):
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            workbook = writer.book
            worksheet = workbook[sheet_name]
            
            warna_header = PatternFill(start_color="1A66CC", end_color="1A66CC", fill_type="solid")
            warna_zebra = PatternFill(start_color="EBF5FF", end_color="EBF5FF", fill_type="solid")
            teks_putih = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            teks_biasa = Font(name="Calibri", size=11, bold=False, color="000000")
            rata_tengah = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.fill = warna_header
                cell.font = teks_putih
                cell.alignment = rata_tengah
            for row_idx in range(2, worksheet.max_row + 1):
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = teks_biasa
                    cell.alignment = rata_tengah
                    if row_idx % 2 == 1:
                        cell.fill = warna_zebra
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                worksheet.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)
        return buffer.getvalue()

    # --- KONSTRUKSI ZONA WAKTU LOKAL INDONESIA ---
    tz_jakarta = pytz.timezone('Asia/Jakarta')

    KAMUS_HARI = {
        "Monday": "Senin", "Tuesday": "Selasa", "Wednesday": "Rabu",
        "Thursday": "Kamis", "Friday": "Jumat", "Saturday": "Sabtu", "Sunday": "Minggu"
    }

    def ambil_hari_ini():
        hari_inggris = datetime.now(tz_jakarta).strftime("%A")
        return KAMUS_HARI.get(hari_inggris, hari_inggris)

    # Ambil Dokumen Spreadsheet Utama
    sh = dapatkan_koneksi_spreadsheet()
    daftar_sheet = [ws.title for ws in sh.worksheets()]
    
    # --- INISIALISASI TAB MASTER PEGAWAI ---
    TAB_MASTER = "Master_Pegawai"
    HEADER_MASTER = ["NIP", "Nama Pegawai", "Jabatan"]
    
    if TAB_MASTER not in daftar_sheet:
        ws_m_init = sh.add_worksheet(title=TAB_MASTER, rows="1000", cols="3")
        ws_m_init.append_row(HEADER_MASTER)
        perbarui_desain_visual_sheet(ws_m_init, jumlah_kolom=3)
        daftar_sheet.append(TAB_MASTER)

    ws_master = sh.worksheet(TAB_MASTER)
    data_master_fisik = ws_master.get_all_values()
    if len(data_master_fisik) == 0:
        ws_master.append_row(HEADER_MASTER)
        perbarui_desain_visual_sheet(ws_master, jumlah_kolom=3)

    try:
        data_master_mentah = ws_master.get_all_records()
        df_pegawai = pd.DataFrame(data_master_mentah)
        if df_pegawai.empty or "NIP" not in df_pegawai.columns:
            df_pegawai = pd.DataFrame(columns=HEADER_MASTER)
    except Exception:
        df_pegawai = pd.DataFrame(columns=HEADER_MASTER)

    if not df_pegawai.empty and "NIP" in df_pegawai.columns:
        df_pegawai.loc[:, "NIP"] = df_pegawai["NIP"].astype(str)

    # Setup Otomatis Nama Sheet Bulanan Aktif Saat Ini
    BULAN_INDO = {
        "January": "Januari", "February": "Februari", "March": "Maret", "April": "April",
        "May": "Mei", "June": "Juni", "July": "Juli", "August": "Agustus",
        "September": "September", "October": "Oktober", "November": "November", "December": "Desember"
    }
    bulan_lokal = BULAN_INDO.get(datetime.now(tz_jakarta).strftime("%B"), datetime.now(tz_jakarta).strftime("%B"))
    sheet_bulan_ini = f"Absen_{bulan_lokal}_{datetime.now(tz_jakarta).strftime('%Y')}"
    
    # 100% SESUAI REALITA GOOGLE SHEETS ANDA (8 KOLOM)
    HEADER_ABSEN = ["Tanggal", "Hari", "NIP", "Nama Pegawai", "Jabatan", "Jam Masuk", "Jam Pulang", "Aktivitas/Pekerjaan Hari Ini"]

    if sheet_bulan_ini not in daftar_sheet:
        ws_b_init = sh.add_worksheet(title=sheet_bulan_ini, rows="1000", cols="8")
        ws_b_init.append_row(HEADER_ABSEN)
        perbarui_desain_visual_sheet(ws_b_init, jumlah_kolom=8)
        daftar_sheet.append(sheet_bulan_ini)

    pilihan_bulan = [t for t in daftar_sheet if t != TAB_MASTER]
    MAX_LOG_BULANAN = 12
    
    if len(pilihan_bulan) > MAX_LOG_BULANAN:
        sheet_tertua_nama = pilihan_bulan[0]
        try:
            ws_dihapus_auto = sh.worksheet(sheet_tertua_nama)
            sh.del_worksheet(ws_dihapus_auto)
            pilihan_bulan.remove(sheet_tertua_nama) 
        except Exception:
            pass

    # --- TAMPILAN INTERFACE UTAMA ---
    st.title("🏢 Office Attendance Cloud")
    
    sheet_aktif = st.selectbox(
        "📂 Pilih Periode / Bulan Absensi:", 
        options=pilihan_bulan, 
        index=pilihan_bulan.index(sheet_bulan_ini) if sheet_bulan_ini in pilihan_bulan else 0
    )
    st.write("---")

    ws_aktif = sh.worksheet(sheet_aktif)
    semua_data = ws_aktif.get_all_records()

    if len(semua_data) == 0:
        df_master = pd.DataFrame(columns=HEADER_ABSEN)
    else:
        df_master = pd.DataFrame(semua_data)
        df_master = df_master[df_master["Tanggal"] != "Tanggal"]
        if "NIP" in df_master.columns:
            df_master.loc[:, "NIP"] = df_master["NIP"].astype(str)

    tab_user, tab_admin = st.tabs(["📝 MODUL ABSENSI KARYAWAN", "🔐 PANEL KONTROL ADMIN"])

    # ==============================================================================
    # TAB UTAMA: MODUL ABSENSI KARYAWAN
    # ==============================================================================
    with tab_user:
        if df_pegawai.empty:
            st.warning("⚠️ Database karyawan kosong. Silakan masuk ke Panel Admin untuk menambah daftar nama karyawan terlebih dahulu.")
        else:
            mode_absen = st.radio("Pilih Tipe Absensi:", ["Absen Masuk (Datang)", "Absen Pulang (Pulang)"], horizontal=True)
            list_dropdown_karyawan = df_pegawai.apply(lambda r: f"{r['NIP']} - {r['Nama Pegawai']}", axis=1).tolist()
            selected_karyawan = st.selectbox("Pilih Nama Anda:", options=list_dropdown_karyawan)
            
            nip_pilihan = selected_karyawan.split(" - ")[0]
            detail_karyawan = df_pegawai[df_pegawai["NIP"] == nip_pilihan].iloc[0]
            
            with st.form("form_absen_karyawan", clear_on_submit=True):
                st.text_input("NIP Pegawai:", value=detail_karyawan["NIP"], disabled=True)
                st.text_input("Nama Lengkap:", value=detail_karyawan["Nama Pegawai"], disabled=True)
                st.text_input("Jabatan / Divisi:", value=detail_karyawan["Jabatan"], disabled=True)
                
                if mode_absen == "Absen Masuk (Datang)":
                    st.info(f"Jam Masuk Anda akan tercatat secara otomatis menggunakan waktu sekarang (WIB).")
                    if st.form_submit_button("🚀 Kirim Absen Masuk", type="primary", use_container_width=True):
                        ws_bulan_sekarang = sh.worksheet(sheet_bulan_ini)
                        
                        waktu_lokal = datetime.now(tz_jakarta)
                        jam_sekarang_str = waktu_lokal.strftime("%H:%M")
                        tanggal_hari_ini = waktu_lokal.strftime("%Y-%m-%d")
                        
                        data_cek = ws_bulan_sekarang.get_all_records()
                        df_cek = pd.DataFrame(data_cek) if data_cek else pd.DataFrame(columns=HEADER_ABSEN)
                        
                        if not df_cek.empty and not df_cek[(df_cek["Tanggal"] == tanggal_hari_ini) & (df_cek["NIP"] == str(nip_pilihan))].empty:
                            st.error(f"❌ Anda ({detail_karyawan['Nama Pegawai']}) sudah melakukan absen masuk hari ini!")
                        else:
                            # 8 Kolom: Tanggal, Hari, NIP, Nama, Jabatan, Jam Masuk, Jam Pulang, Aktivitas
                            baris_absen = [tanggal_hari_ini, ambil_hari_ini(), detail_karyawan["NIP"], detail_karyawan["Nama Pegawai"], detail_karyawan["Jabatan"], jam_sekarang_str, "-", "-"]
                            ws_bulan_sekarang.append_row(baris_absen)
                            perbarui_desain_visual_sheet(ws_bulan_sekarang, jumlah_kolom=8)
                            st.success(f"🎉 Absen masuk disimpan jam {jam_sekarang_str} WIB.")
                            st.rerun()
                            
                else: 
                    pekerjaan_hari_ini = st.text_area("Laporan Aktivitas / Realisasi Pekerjaan Hari Ini:", placeholder="Tuliskan pekerjaan Anda...")
                    if st.form_submit_button("🔒 Kirim Absen Pulang", type="primary", use_container_width=True):
                        if not pekerjaan_hari_ini or pekerjaan_hari_ini.strip() == "":
                            st.error("⚠️ Gagal! Anda wajib mengisi laporan aktivitas sebelum melakukan absen pulang.")
                        else:
                            ws_bulan_sekarang = sh.worksheet(sheet_bulan_ini)
                            waktu_lokal = datetime.now(tz_jakarta)
                            tanggal_hari_ini = waktu_lokal.strftime("%Y-%m-%d")
                            jam_pulang_str = waktu_lokal.strftime("%H:%M")
                            data_cek = ws_bulan_sekarang.get_all_records()
                            baris_ketemu = False
                            
                            if data_cek:
                                for idx, row in enumerate(data_cek):
                                    if str(row.get("Tanggal")) == tanggal_hari_ini and str(row.get("NIP")) == nip_pilihan:
                                        baris_sheet = idx + 2
                                        ws_bulan_sekarang.update_cell(baris_sheet, 7, jam_pulang_str)       # Kolom G (Jam Pulang)
                                        ws_bulan_sekarang.update_cell(baris_sheet, 8, pekerjaan_hari_ini)   # Kolom H (Aktivitas)
                                        baris_ketemu = True
                                        break
                                    
                            if baris_ketemu:
                                perbarui_desain_visual_sheet(ws_bulan_sekarang, jumlah_kolom=8)
                                st.success(f"✨ Absen pulang tercatat jam {jam_pulang_str} WIB.")
                                st.rerun()
                            else:
                                st.error(f"❌ Log absen masuk tanggal {tanggal_hari_ini} tidak ditemukan di bulan ini.")

        if not df_master.empty:
            st.write("---")
            st.subheader(f"📊 Manifes Log Absensi - Periode {sheet_aktif}")
            st.dataframe(df_master, use_container_width=True, hide_index=True)
            excel_data_pegawai = buat_excel_bytes(df_master, sheet_aktif)
            st.download_button(
                label=f"📥 Unduh Log Manifes Absen Berformat Excel ({sheet_aktif}.xlsx)",
                data=excel_data_pegawai,
                file_name=f"Log_Absensi_{sheet_aktif}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    # ==============================================================================
    # TAB KEDUA: PANEL KONTROL ADMIN
    # ==============================================================================
    with tab_admin:
        st.subheader("🔒 Verifikasi Autentikasi Pengelola")
        key_admin = st.text_input("Masukkan Kode Akses Admin Kantor:", type="password")
        
        if key_admin == "admin123":
            st.success("Akses Terbuka. Silakan kelola manajemen kantor di bawah ini.")
            st.write("---")
            
            st.markdown("### 📊 📈 Rekapitulasi Performa Bulanan (Laporan Atasan)")
            if df_master.empty:
                st.info(f"Belum ada data absensi di tab periode '{sheet_aktif}' untuk dihitung.")
            else:
                rekap_hadir = df_master.groupby(['NIP', 'Nama Pegawai', 'Jabatan']).size().reset_index(name='Total Hari Kerja')
                st.dataframe(rekap_hadir, use_container_width=True, hide_index=True)
                
                buffer_boss = io.BytesIO()
                with pd.ExcelWriter(buffer_boss, engine='openpyxl') as writer:
                    rekap_hadir.to_excel(writer, index=False, sheet_name="Ringkasan_Absensi")
                    worksheet = writer.book["Ringkasan_Absensi"]
                    warna_header = PatternFill(start_color="1E4620", end_color="1E4620", fill_type="solid")
                    warna_zebra = PatternFill(start_color="F1F8F1", end_color="F1F8F1", fill_type="solid")
                    teks_putih = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                    teks_biasa = Font(name="Arial", size=11, bold=False, color="000000")
                    rata_tengah = Alignment(horizontal="center", vertical="center")
                    
                    for cell in worksheet[1]:
                        cell.fill = warna_header
                        cell.font = teks_putih
                        cell.alignment = rata_tengah
                    for row_idx in range(2, worksheet.max_row + 1):
                        for col_idx in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.font = teks_biasa
                            cell.alignment = rata_tengah
                            if row_idx % 2 == 1: cell.fill = warna_zebra
                    for col in worksheet.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        worksheet.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 15)
                
                st.download_button(
                    label="🟢 Unduh Berkas Summary Ringkasan Excel (Untuk Atasan)",
                    data=buffer_boss.getvalue(),
                    file_name=f"SUMMARY_ABSENSI_{sheet_aktif}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            st.write("---")
            st.markdown("### 📥 Ekspor Manifes Penuh (Unduh Versi Admin)")
            if not df_master.empty:
                excel_data_admin = buat_excel_bytes(df_master, sheet_aktif)
                st.download_button(
                    label=f"📥 Download Berkas Log Detail Penuh ({sheet_aktif}.xlsx)",
                    data=excel_data_admin,
                    file_name=f"Detail_Log_{sheet_aktif}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_admin_pilih"
                )
            
            st.write("---")
            st.markdown("### ⚙️ Pengaturan Profil Karyawan & Data Master")
            
            with st.expander("➕ Tambahkan Profil Karyawan Baru"):
                with st.form("form_tambah_p", clear_on_submit=True):
                    add_nip = st.text_input("NIP Baru (Harus Unik):")
                    add_nama = st.text_input("Nama Lengkap:")
                    add_jabatan = st.text_input("Jabatan / Posisi Kerja:")
                    if st.form_submit_button("Simpan Karyawan ke Database"):
                        if add_nip and add_nama and add_jabatan:
                            if add_nip in df_pegawai["NIP"].tolist():
                                st.error("❌ Gagal! NIP sudah terdaftar di database.")
                            else:
                                ws_master.append_row([str(add_nip), add_nama, add_jabatan])
                                perbarui_desain_visual_sheet(ws_master, jumlah_kolom=3)
                                st.success(f"🎉 Pegawai baru {add_nama} berhasil didaftarkan!")
                                st.rerun()
                        else:
                            st.warning("⚠️ Semua data wajib diisi!")

            with st.expander("📝 Edit / Perbarui Data Karyawan"):
                if df_pegawai.empty:
                    st.info("Database kosong.")
                else:
                    karyawan_diedit = st.selectbox("Pilih Karyawan yang Ingin Diperbarui:", options=list_dropdown_karyawan, key="sb_edit")
                    nip_edit_target = karyawan_diedit.split(" - ")[0]
                    idx_pegawai = df_pegawai[df_pegawai["NIP"] == nip_edit_target].index[0]
                    row_data_lama = df_pegawai.loc[idx_pegawai]
                    with st.form("form_edit_p"):
                        new_nip = st.text_input("Perbarui NIP:", value=str(row_data_lama["NIP"]))
                        new_nama = st.text_input("Perbarui Nama Lengkap:", value=row_data_lama["Nama Pegawai"])
                        new_jabatan = st.text_input("Perbarui Jabatan / Divisi:", value=row_data_lama["Jabatan"])
                        if st.form_submit_button("Simpan Perubahan Data"):
                            baris_sheet = int(idx_pegawai) + 2
                            ws_master.update_cell(baris_sheet, 1, str(new_nip))
                            ws_master.update_cell(baris_sheet, 2, new_nama)
                            ws_master.update_cell(baris_sheet, 3, new_jabatan)
                            perbarui_desain_visual_sheet(ws_master, jumlah_kolom=3)
                            st.success("🎉 Perubahan data karyawan berhasil disinkronkan!")
                            st.rerun()

            with st.expander("🗑️ Hapus Akun Karyawan dari Master"):
                if df_pegawai.empty:
                    st.info("Database master kosong.")
                else:
                    karyawan_dihapus = st.selectbox("Pilih Karyawan yang Ingin Dihapus:", options=list_dropdown_karyawan, key="sb_hapus")
                    nip_hapus_target = karyawan_dihapus.split(" - ")[0]
                    if st.button("Hapus Permanen Karyawan", type="primary", use_container_width=True):
                        idx_hapus = df_pegawai[df_pegawai["NIP"] == nip_hapus_target].index[0]
                        ws_master.delete_rows(int(idx_hapus) + 2)
                        perbarui_desain_visual_sheet(ws_master, jumlah_kolom=3)
                        st.success("🗑️ Data pegawai berhasil dihapus dari Master!")
                        st.rerun()

            st.write("---")
            st.subheader("⚙️ Manajer Penghapusan Data Log Kontrol")
            
            sheet_target_operasi = st.selectbox("Pilih Target Nama Sheet Bulanan:", options=pilihan_bulan, key="sb_operasi_admin")
            ws_target_ops = sh.worksheet(sheet_target_operasi)
            
            if st.button(f"⚠️ Hapus 1 Baris Data Log Terakhir pada Sheet: {sheet_target_operasi}", use_container_width=True):
                try:
                    nilai_fisik_sheet = ws_target_ops.get_all_values()
                    total_baris_fisik = len(nilai_fisik_sheet)
                    if total_baris_fisik > 1:
                        ws_target_ops.delete_rows(total_baris_fisik)
                        perbarui_desain_visual_sheet(ws_target_ops, jumlah_kolom=8)
                        st.success(f"✔️ Sukses! Satu baris transaksi terakhir telah terhapus.")
                        st.rerun()
                    else:
                        st.warning(f"ℹ️ Lembar sudah bersih kosong.")
                except Exception as e:
                    st.error(f"Gagal menghapus baris data: {e}")

            st.write("")
            st.markdown("---")
            st.markdown("### 🚨 Opsi Reset Total (Wipe Out Database)")
            
            konfirmasi_wipe_out = st.checkbox("Saya sadar dan setuju untuk menghapus semua sheet bulanan.")
            
            if st.button("🚨 RESTART & WIPE OUT: BERSIHKAN SEMUA DATA SHEET LOG", type="primary", use_container_width=True, disabled=not konfirmasi_wipe_out):
                try:
                    sheet_pemulihan_sementara = f"Mulai_Baru_{sheet_bulan_ini}"
                    ws_temp = sh.add_worksheet(title=sheet_pemulihan_sementara, rows="1000", cols="8")
                    ws_temp.append_row(HEADER_ABSEN)
                    
                    for sheet_loop in sh.worksheets():
                        if sheet_loop.title != sheet_pemulihan_sementara and sheet_loop.title != TAB_MASTER:
                            sh.del_worksheet(sheet_loop)
                    
                    ws_temp.update_title(sheet_bulan_ini)
                    perbarui_desain_visual_sheet(ws_temp, jumlah_kolom=8)
                    
                    st.success("💥 Database Berhasil Direset Total!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal mengeksekusi sistem wipe-out: {e}")
                    
        elif key_admin != "":
            st.error("❌ Kode Akses Admin Salah!")
